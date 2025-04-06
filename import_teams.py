from app import app, db, Team, User
from openpyxl import load_workbook

def parse_cell_content(content):
    if not content:
        return {}
    
    lines = str(content).split('\n')
    data = {}
    for line in lines:
        if ':' in line:
            key, value = line.split(':', 1)
            data[key.strip()] = value.strip()
    return data

def import_teams_from_excel():
    try:
        print("Excel dosyası okunuyor...")
        wb = load_workbook('hiyerarsi.xlsx')
        ws = wb['Sayfa2']
        
        # Önce tüm üst ekip gruplarını bulalım
        parent_types = set()
        team_categories = set()
        
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    data = parse_cell_content(cell.value)
                    if 'Üst Ekip' in data:
                        parent_types.add(data['Üst Ekip'])
                    if 'Ekip' in data:
                        team_categories.add(data['Ekip'])
        
        print("\nBulunan Üst Ekipler:")
        for pt in parent_types:
            print(f"- {pt}")
            
        print("\nBulunan Ekip Kategorileri:")
        for tc in team_categories:
            print(f"- {tc}")
        
        # Önce üst ekip gruplarını oluşturalım
        teams = {}  # name: Team object
        for parent_type in parent_types:
            team = Team(
                name=parent_type,
                description=f"{parent_type} üst ekibi",
                parent_type=parent_type,
                team_category=parent_type
            )
            db.session.add(team)
            teams[parent_type] = team
            print(f"Üst ekip oluşturuldu: {parent_type}")
        
        # Şimdi çalışma birimlerini oluşturalım
        for category in team_categories:
            if category not in teams:  # Eğer üst ekip olarak zaten oluşturulmadıysa
                # İlgili üst ekibi bulalım
                parent_type = None
                for row in range(1, ws.max_row + 1):
                    for col in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row, column=col)
                        if cell.value:
                            data = parse_cell_content(cell.value)
                            if data.get('Ekip') == category and 'Üst Ekip' in data:
                                parent_type = data['Üst Ekip']
                                break
                    if parent_type:
                        break
                
                team = Team(
                    name=category,
                    description=f"{category} ekibi",
                    parent_type=parent_type if parent_type else "Diğer",
                    team_category=category
                )
                if parent_type and parent_type in teams:
                    team.parent = teams[parent_type]
                
                db.session.add(team)
                teams[category] = team
                print(f"Çalışma birimi oluşturuldu: {category} (Üst Ekip: {parent_type})")
        
        db.session.commit()
        print("\nEkip yapısı başarıyla oluşturuldu!")
        
        # Oluşturulan ekipleri göster
        all_teams = Team.query.all()
        print(f"\nToplam {len(all_teams)} ekip oluşturuldu:")
        for team in all_teams:
            parent_name = team.parent.name if team.parent else "Yok"
            print(f"Ekip: {team.name}")
            print(f"  Üst Ekip Türü: {team.parent_type}")
            print(f"  Bağlı Olduğu Ekip: {parent_name}")
            print(f"  Kategori: {team.team_category}")
            print("-" * 40)
            
    except Exception as e:
        print(f"Hata oluştu: {str(e)}")
        import traceback
        print("Hata detayı:")
        print(traceback.format_exc())
        db.session.rollback()

def import_teams():
    with app.app_context():
        # Ana ekipler ve kategorileri
        parent_teams = {
            "Dış Giyim P&B": "Dış Giyim P&B",
            "Tasarım Sorumlusu": "Tasarım Sorumlusu",
            "Örme Key T-Shirt / Hırka P&B": "Örme Key T-Shirt / Hırka P&B",
            "Triko P&B": "Triko P&B",
            "BC & CPPM & SM": "BC & CPPM & SM",
            "Dokuma Üst P&B": "Dokuma Üst P&B",
            "Design Team": "Design Team",
            "Dokuma Alt P&B": "Dokuma Alt P&B",
            "Direktör Altı Pozisyonlar": "Direktör Altı Pozisyonlar",
            "Baskılı T-Shirt / Sweat P&B": "Baskılı T-Shirt / Sweat P&B",
            "PBL'ler": "PBL'ler",
            "Garment": "Garment"
        }

        # Alt ekipler ve bağlı oldukları üst ekipler
        sub_teams = {
            "BC": ("Direktör Altı Pozisyonlar", "BC"),
            "Dış Giyim Buying": ("Dış Giyim P&B", "Dış Giyim Buying"),
            "YD CPPM": ("BC & CPPM & SM", "YD CPPM"),
            "Örme Baskılı T-Shirt / Sweat Design": ("Design Team", "Örme Baskılı T-Shirt / Sweat Design"),
            "Tasarım İletişim": ("Tasarım Sorumlusu", "Tasarım İletişim"),
            "Baskılı T-Shirt / Sweat Buying": ("Baskılı T-Shirt / Sweat P&B", "Baskılı T-Shirt / Sweat Buying"),
            "Dokuma Üst Design": ("Design Team", "Dokuma Üst Design"),
            "Dokuma Üst Buying": ("Dokuma Üst P&B", "Dokuma Üst Buying"),
            "Örme Key T-Shirt / Hırka Buying": ("Örme Key T-Shirt / Hırka P&B", "Örme Key T-Shirt / Hırka Buying"),
            "Triko Design": ("Design Team", "Triko Design"),
            "Örme Key T-Shirt / Hırka Design": ("Design Team", "Örme Key T-Shirt / Hırka Design"),
            "Distrubiton": ("Direktör Altı Pozisyonlar", "Distrubiton"),
            "Triko Buying": ("Triko P&B", "Triko Buying"),
            "SMM": ("BC & CPPM & SM", "SMM"),
            "Dokuma Alt Buying": ("Dokuma Alt P&B", "Dokuma Alt Buying"),
            "Dokuma Alt Design": ("Design Team", "Dokuma Alt Design"),
            "Kategori Direktörleri": ("Direktör Altı Pozisyonlar", "Kategori Direktörleri")
        }

        try:
            # Önce ana ekipleri oluştur
            for team_name, category in parent_teams.items():
                team = Team(
                    name=team_name,
                    parent_type=team_name,
                    team_category=category,
                    description=f"{team_name} ekibi"
                )
                db.session.add(team)
                print(f"Üst ekip oluşturuldu: {team_name}")
            
            db.session.commit()

            # Sonra alt ekipleri oluştur
            for team_name, (parent_name, category) in sub_teams.items():
                parent_team = Team.query.filter_by(name=parent_name).first()
                if parent_team:
                    team = Team(
                        name=team_name,
                        parent_type=parent_name,
                        parent_id=parent_team.id,
                        team_category=category,
                        description=f"{team_name} ekibi"
                    )
                    db.session.add(team)
                    print(f"Alt ekip oluşturuldu: {team_name} (Üst Ekip: {parent_name})")
            
            db.session.commit()
            print("\nTüm ekipler başarıyla oluşturuldu!")

        except Exception as e:
            db.session.rollback()
            print(f"Hata oluştu: {str(e)}")

if __name__ == "__main__":
    import_teams() 