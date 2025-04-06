import pandas as pd
from openpyxl import load_workbook

def parse_cell_content(content):
    if not content:
        return {}
    
    lines = str(content).split('\n')
    data = {}
    for line in lines:
        if ':' in line:
            key, value = line.split(':', 1)
            data[key.strip()] = value.strip()
    return data

print("Excel dosyası okunuyor...")
wb = load_workbook('hiyerarsi.xlsx')
ws = wb['Sayfa2']

print("\nHücre içerikleri:")
print("-" * 80)

for row in range(1, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        if cell.value:
            data = parse_cell_content(cell.value)
            if data:
                print(f"\nHücre [{row},{col}]:")
                for key, value in data.items():
                    print(f"{key}: {value}")
                print("-" * 40)

for sheet_name in wb.sheetnames:
    print(f"\n{sheet_name} sayfası:")
    ws = wb[sheet_name]
    print(f"Satır sayısı: {ws.max_row}")
    print(f"Sütun sayısı: {ws.max_column}")
    
    print("\nİlk 5x5 hücre değerleri:")
    for row in range(1, min(6, ws.max_row + 1)):
        row_values = []
        for col in range(1, min(6, ws.max_column + 1)):
            cell = ws.cell(row=row, column=col)
            row_values.append(str(cell.value if cell.value is not None else ''))
        print(f"Satır {row}: {row_values}") 